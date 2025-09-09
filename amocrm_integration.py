# amocrm_integration.py
from __future__ import annotations

import base64
import hashlib
import hmac
import io
import json
import os
import time
from typing import Any, Dict, Optional, Tuple
from collections import defaultdict

import requests
from flask import (
    Blueprint,
    current_app,
    jsonify,
    redirect,
    render_template,
    request,
    session,
    send_file,
)

# ======== DB ========
try:
    from extensions import db  # type: ignore
except Exception:
    db = None  # type: ignore

try:
    from models import Company  # type: ignore
except Exception:
    Company = None  # type: ignore


# ======== Optional Excel ========
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except Exception:
    Workbook = None  # type: ignore


# ======== ENV ========
AMO_CLIENT_ID = os.getenv("AMO_CLIENT_ID", "")
AMO_CLIENT_SECRET = os.getenv("AMO_CLIENT_SECRET", "")
AMO_REDIRECT_BASE = os.getenv("AMO_REDIRECT_BASE", "http://localhost:5000")
AMO_STATE_SECRET = os.getenv("AMO_STATE_SECRET", "change-me")

WON_STATUS_ID = 142
LOST_STATUS_ID = 143


# ======== Blueprints ========
bp_amocrm_company_api = Blueprint(
    "amocrm_company_api",
    __name__,
    url_prefix="/api/partners/company",
)
bp_amocrm_pages = Blueprint(
    "amocrm_pages",
    __name__,
    url_prefix="/partner/company",
)


# ======== Utils ========
def _b64url(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).decode().rstrip("=")

def _sign_state(payload: dict) -> str:
    raw = json.dumps(payload, separators=(",", ":"), ensure_ascii=False).encode()
    sig = hmac.new(AMO_STATE_SECRET.encode(), raw, hashlib.sha256).digest()
    return _b64url(sig)

def _get_company(company_id: int):
    if Company and db:
        return db.session.get(Company, company_id)
    return None

def _session_key(company_id: int) -> str:
    return f"amo_tokens:{company_id}"

def _save_tokens(company_id: int, data: Dict[str, Any]) -> None:
    expires_at = int(time.time()) + int(data.get("expires_in", 0))
    base_domain = data.get("base_domain") or data.get("domain")
    if Company and db:
        company = _get_company(company_id)
        if company is None:
            return
        setattr(company, "amo_access_token", data.get("access_token"))
        setattr(company, "amo_refresh_token", data.get("refresh_token"))
        setattr(company, "amo_token_type", data.get("token_type"))
        setattr(company, "amo_domain", base_domain)
        setattr(company, "amo_expires_at", expires_at)
        if hasattr(company, "amo_last_sync_at"):
            setattr(company, "amo_last_sync_at", int(time.time()))
        db.session.add(company)
        db.session.commit()
    else:
        session[_session_key(company_id)] = {
            "access_token": data.get("access_token"),
            "refresh_token": data.get("refresh_token"),
            "token_type": data.get("token_type"),
            "base_domain": base_domain,
            "expires_at": expires_at,
            "last_sync_at": int(time.time()),
        }

def _read_tokens(company_id: int) -> Optional[Dict[str, Any]]:
    if Company and db:
        c = _get_company(company_id)
        if not c:
            return None
        access = getattr(c, "amo_access_token", None)
        if not access:
            return None
        return {
            "access_token": access,
            "refresh_token": getattr(c, "amo_refresh_token", None),
            "token_type": getattr(c, "amo_token_type", None),
            "base_domain": getattr(c, "amo_domain", None),
            "expires_at": getattr(c, "amo_expires_at", None),
            "last_sync_at": getattr(c, "amo_last_sync_at", None) if hasattr(c, "amo_last_sync_at") else None,
        }
    return session.get(_session_key(company_id))

def _clear_tokens(company_id: int) -> None:
    if Company and db:
        c = _get_company(company_id)
        if c:
            for k in ("amo_access_token", "amo_refresh_token", "amo_token_type", "amo_domain", "amo_expires_at", "amo_last_sync_at"):
                if hasattr(c, k):
                    setattr(c, k, None)
            db.session.add(c)
            db.session.commit()
    session.pop(_session_key(company_id), None)

def _callback_url(company_id: int) -> str:
    return f"{AMO_REDIRECT_BASE}/api/partners/company/{company_id}/crm/amocrm/callback"

def _partner_company_url(company_id: int) -> str:
    return f"/partner/company/{company_id}/crm"

def _amo_headers(access_token: str) -> Dict[str, str]:
    return {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}

def _platform_host_from_domain(dom: str | None) -> str:
    d = (dom or "").lower()
    if d.endswith(".amocrm.ru") or ".amocrm.ru" in d or d == "":
        return "www.amocrm.ru"
    return "www.amocrm.com"

def _refresh_if_needed(company_id: int) -> Optional[Dict[str, Any]]:
    t = _read_tokens(company_id)
    if not t:
        return None
    exp = int(t.get("expires_at") or 0)
    now = int(time.time())
    if now + 60 < exp:
        return t
    refresh = t.get("refresh_token")
    base_domain = t.get("base_domain")
    if not (refresh and base_domain and AMO_CLIENT_ID and AMO_CLIENT_SECRET):
        return t
    try:
        token_url = f"https://{base_domain}/oauth2/access_token"
        payload = {
            "client_id": AMO_CLIENT_ID,
            "client_secret": AMO_CLIENT_SECRET,
            "grant_type": "refresh_token",
            "refresh_token": refresh,
            "redirect_uri": _callback_url(company_id),
        }
        r = requests.post(token_url, json=payload, timeout=15)
        if r.status_code == 200:
            data = r.json()
            if "base_domain" not in data:
                data["base_domain"] = base_domain
            _save_tokens(company_id, data)
            return _read_tokens(company_id)
        else:
            current_app.logger.warning("AMO refresh failed: %s %s", r.status_code, r.text)
            return t
    except requests.RequestException:
        current_app.logger.exception("AMO refresh exception")
        return t

def _amo_get(base_domain: str, access_token: str, path: str, params: Dict[str, Any]) -> requests.Response:
    url = f"https://{base_domain}{path}"
    return requests.get(url, headers=_amo_headers(access_token), params=params, timeout=20)

def _fetch_users_map(base_domain: str, access_token: str) -> Dict[int, Dict[str, Any]]:
    """
    Возвращает { id: {id, name, email?} }
    """
    out: Dict[int, Dict[str, Any]] = {}
    page, limit = 1, 250
    while True:
        r = _amo_get(base_domain, access_token, "/api/v4/users", {"page": page, "limit": limit})
        if r.status_code != 200:
            current_app.logger.warning("AMO users fetch %s: %s", r.status_code, r.text)
            break
        data = r.json() or {}
        users = (data.get("_embedded") or {}).get("users") or []
        for u in users:
            uid = u.get("id")
            if uid is not None:
                out[int(uid)] = {
                    "id": int(uid),
                    "name": u.get("name") or f"User {uid}",
                    "email": (u.get("email") or "") if isinstance(u.get("email"), str) else "",
                }
        if len(users) < limit:
            break
        page += 1
    return out

def _iter_closed_leads(base_domain: str, access_token: str, ts_from: int, ts_to: int):
    page, limit = 1, 250
    while True:
        params = {
            "page": page,
            "limit": limit,
            "filter[closed_at][from]": ts_from,
            "filter[closed_at][to]": ts_to,
            "order[closed_at]": "desc",
        }
        r = _amo_get(base_domain, access_token, "/api/v4/leads", params)
        if r.status_code != 200:
            current_app.logger.error("AMO leads fetch %s: %s", r.status_code, r.text)
            break
        data = r.json() or {}
        leads = (data.get("_embedded") or {}).get("leads") or []
        if not leads:
            break
        for lead in leads:
            yield lead
        if len(leads) < limit:
            break
        page += 1

def _iter_created_leads(base_domain: str, access_token: str, ts_from: int, ts_to: int):
    page, limit = 1, 250
    while True:
        params = {
            "page": page,
            "limit": limit,
            "filter[created_at][from]": ts_from,
            "filter[created_at][to]": ts_to,
            "order[created_at]": "desc",
        }
        r = _amo_get(base_domain, access_token, "/api/v4/leads", params)
        if r.status_code != 200:
            current_app.logger.error("AMO leads(created) fetch %s: %s", r.status_code, r.text)
            break
        data = r.json() or {}
        leads = (data.get("_embedded") or {}).get("leads") or []
        if not leads:
            break
        for lead in leads:
            yield lead
        if len(leads) < limit:
            break
        page += 1

def _period_from_request() -> Tuple[int, int, int, str]:
    now = int(time.time())
    rng = (request.args.get("range") or "").lower().strip()
    a_from = request.args.get("from")
    a_to = request.args.get("to")
    a_days = request.args.get("days")

    if rng == "today" or a_days == "0":
        lt = time.localtime(now)
        midnight = int(time.mktime((lt.tm_year, lt.tm_mon, lt.tm_mday, 0, 0, 0, lt.tm_wday, lt.tm_yday, lt.tm_isdst)))
        return midnight, now, 0, "today"

    if a_from and a_to:
        try:
            ts_from = int(a_from)
            ts_to = int(a_to)
            days = max(0, int((ts_to - ts_from) / 86400))
            return ts_from, ts_to, days, "custom"
        except Exception:
            pass

    try:
        days = int(a_days) if a_days is not None else 30
    except Exception:
        days = 30
    ts_to = now
    ts_from = now - days * 86400
    label = f"{days}d"
    return ts_from, ts_to, days, label

def _compute_stats(base_domain: str, access_token: str, ts_from: int, ts_to: int) -> Dict[str, Any]:
    users_map = _fetch_users_map(base_domain, access_token)  # {id:{id,name,email}}
    by_user = defaultdict(lambda: {"won": 0, "lost": 0})
    total_won = total_lost = 0

    for lead in _iter_closed_leads(base_domain, access_token, ts_from, ts_to):
        status_id = int(lead.get("status_id") or 0)
        resp_id = lead.get("responsible_user_id") or 0
        if status_id == WON_STATUS_ID:
            by_user[resp_id]["won"] += 1
            total_won += 1
        elif status_id == LOST_STATUS_ID:
            by_user[resp_id]["lost"] += 1
            total_lost += 1

    rows = []
    for uid, agg in by_user.items():
        won, lost = agg["won"], agg["lost"]
        total = won + lost
        conv = round(100 * won / total) if total else 0
        name = (users_map.get(uid) or {}).get("name") if uid else "Без владельца"
        rows.append({"user_id": uid, "display_name": name, "won": won, "lost": lost, "conv": conv})

    return {
        "won_count": total_won,
        "lost_count": total_lost,
        "rows": rows,
    }

def _apply_view_filters(rows: list[Dict[str, Any]], sort: str, min_total: int, q: str) -> list[Dict[str, Any]]:
    if q:
        ql = q.lower()
        rows = [r for r in rows if (r.get("display_name") or "").lower().find(ql) >= 0]
    if min_total > 0:
        rows = [r for r in rows if (r.get("won", 0) + r.get("lost", 0)) >= min_total]
    if sort == "conv_desc":
        rows.sort(key=lambda r: (r.get("conv", 0), r.get("won", 0)), reverse=True)
    elif sort == "lost_asc":
        rows.sort(key=lambda r: (r.get("lost", 0), -r.get("won", 0)))
    elif sort == "name_asc":
        rows.sort(key=lambda r: str(r.get("display_name") or ""))
    else:  # won_desc
        rows.sort(key=lambda r: (-r.get("won", 0), r.get("lost", 0)))
    return rows


# ======== Persistent Mapping (DB + fallback to session) ========

AmoUserLink = None
if db:
    try:
        from sqlalchemy import UniqueConstraint

        class AmoUserLink(db.Model):  # type: ignore
            __tablename__ = "amo_user_link"
            id = db.Column(db.Integer, primary_key=True)
            company_id = db.Column(db.Integer, index=True, nullable=False)
            platform_user_id = db.Column(db.Integer, nullable=False)
            amocrm_user_id = db.Column(db.Integer, nullable=False)
            updated_at = db.Column(db.Integer, nullable=False, default=lambda: int(time.time()))
            __table_args__ = (UniqueConstraint("company_id", "platform_user_id", name="uq_amo_user_link"),)
    except Exception:
        AmoUserLink = None  # type: ignore

def _ensure_link_table():
    if not (db and AmoUserLink):
        return False
    try:
        engine = db.engine
        if not engine.has_table("amo_user_link"):  # type: ignore
            AmoUserLink.__table__.create(bind=engine)  # type: ignore
        return True
    except Exception:
        return False

def _get_mapping_db(company_id: int) -> Dict[str, int]:
    if not (db and AmoUserLink and _ensure_link_table()):
        return {}
    rows = AmoUserLink.query.filter_by(company_id=company_id).all()  # type: ignore
    m: Dict[str, int] = {}
    for r in rows:
        m[str(r.platform_user_id)] = int(r.amocrm_user_id)
    return m

def _set_mapping_db(company_id: int, platform_user_id: int, amocrm_user_id: int) -> None:
    if not (db and AmoUserLink and _ensure_link_table()):
        return
    row = AmoUserLink.query.filter_by(company_id=company_id, platform_user_id=platform_user_id).first()  # type: ignore
    now = int(time.time())
    if row:
        row.amocrm_user_id = amocrm_user_id
        row.updated_at = now
        db.session.add(row)
    else:
        db.session.add(AmoUserLink(
            company_id=company_id,
            platform_user_id=platform_user_id,
            amocrm_user_id=amocrm_user_id,
            updated_at=now
        ))
    db.session.commit()

def _get_mapping(company_id: int) -> Dict[str, int]:
    if db and AmoUserLink and _ensure_link_table():
        m = _get_mapping_db(company_id)
        if m:
            return m
    # fallback
    return session.get(f"amo_user_map:{company_id}") or {}

def _set_mapping(company_id: int, platform_user_id: int, amocrm_user_id: int) -> None:
    # persist in DB if possible
    try:
        _set_mapping_db(company_id, platform_user_id, amocrm_user_id)
    except Exception:
        pass
    # and keep session mirror
    key = f"amo_user_map:{company_id}"
    m = session.get(key) or {}
    m[str(platform_user_id)] = int(amocrm_user_id)
    session[key] = m


# ======== API ========

@bp_amocrm_company_api.get("/<int:company_id>/crm/amocrm/status")
def amocrm_status(company_id: int):
    tok = _refresh_if_needed(company_id)
    if not tok:
        return jsonify({"connected": False})
    return jsonify({
        "connected": True,
        "base_domain": tok.get("base_domain"),
        "token_expires_at": tok.get("expires_at"),
        "last_sync_at": tok.get("last_sync_at"),
    })

@bp_amocrm_company_api.post("/<int:company_id>/crm/amocrm/unlink")
def amocrm_unlink(company_id: int):
    _clear_tokens(company_id)
    return jsonify({"ok": True})

@bp_amocrm_company_api.post("/<int:company_id>/crm/amocrm/connect")
def amocrm_connect(company_id: int):
    typed_domain = (request.form.get("base_domain") or
                    (request.json.get("base_domain") if request.is_json else None) or "").strip()

    if not (AMO_CLIENT_ID and AMO_CLIENT_SECRET and AMO_REDIRECT_BASE):
        return jsonify({"error": "AMO_* переменные окружения не заданы"}), 500

    payload = {"cid": company_id, "ts": int(time.time())}
    state = _b64url(json.dumps({"p": payload, "s": _sign_state(payload)}, separators=(",", ":")).encode())
    platform_host = _platform_host_from_domain(typed_domain or "amocrm.ru")
    query = f"client_id={requests.utils.quote(AMO_CLIENT_ID)}&state={requests.utils.quote(state)}&mode=post_message"
    auth_url = f"https://{platform_host}/oauth?{query}"

    if typed_domain:
        session[f"amo_auth_domain:{company_id}"] = typed_domain

    current_app.logger.warning("AMO auth_url -> %s", auth_url)

    accept = request.headers.get("Accept", "")
    if "application/json" in accept or request.is_json:
        return jsonify({"auth_url": auth_url})
    return redirect(auth_url, code=302)

@bp_amocrm_company_api.get("/<int:company_id>/crm/amocrm/callback")
def amocrm_callback(company_id: int):
    code = request.args.get("code")
    state = request.args.get("state")
    referer_dom = (request.args.get("referer") or "").strip()
    session_dom = session.get(f"amo_auth_domain:{company_id}")
    dom = referer_dom or session_dom

    if not code or not state or not dom:
        return "Invalid OAuth callback", 400

    try:
        raw = base64.urlsafe_b64decode(state + "==")
        parsed = json.loads(raw.decode())
        payload = parsed.get("p") or {}
        sig = parsed.get("s") or ""
        if _sign_state(payload) != sig or int(payload.get("cid")) != int(company_id):
            return "Bad state", 400
    except Exception:
        return "Bad state", 400

    token_url = f"https://{dom}/oauth2/access_token"
    data = {
        "client_id": AMO_CLIENT_ID,
        "client_secret": AMO_CLIENT_SECRET,
        "grant_type": "authorization_code",
        "code": code,
        "redirect_uri": _callback_url(company_id),
    }
    try:
        r = requests.post(token_url, json=data, timeout=15)
        if r.status_code != 200:
            current_app.logger.error("AMO token exchange failed: %s %s", r.status_code, r.text)
            return "Token exchange failed", 400
        tok = r.json()
        if "base_domain" not in tok:
            tok["base_domain"] = dom
        _save_tokens(company_id, tok)
    except requests.RequestException:
        current_app.logger.exception("AMO token exchange error")
        return "Token request error", 500

    return redirect(_partner_company_url(company_id), code=302)

@bp_amocrm_company_api.post("/<int:company_id>/crm/amocrm/sync")
def amocrm_sync(company_id: int):
    tok = _read_tokens(company_id)
    if not tok:
        return jsonify({"error": "not connected"}), 400
    if Company and db:
        c = _get_company(company_id)
        if c and hasattr(c, "amo_last_sync_at"):
            setattr(c, "amo_last_sync_at", int(time.time()))
            db.session.add(c)
            db.session.commit()
    else:
        tok["last_sync_at"] = int(time.time())
        session[_session_key(company_id)] = tok
    return jsonify({"ok": True})

@bp_amocrm_company_api.get("/<int:company_id>/crm/stats")
def crm_stats(company_id: int):
    tok = _refresh_if_needed(company_id)
    if not tok:
        return jsonify({"error": "not connected"}), 400

    base_domain = tok.get("base_domain")
    access_token = tok.get("access_token")
    if not base_domain or not access_token:
        return jsonify({"error": "not connected"}), 400

    ts_from, ts_to, days, label = _period_from_request()
    data = _compute_stats(base_domain, access_token, ts_from, ts_to)

    total_all = data["won_count"] + data["lost_count"]
    conversion = round(100 * data["won_count"] / total_all) if total_all else 0

    return jsonify({
        "range": label,
        "from": ts_from,
        "to": ts_to,
        "days": days,
        "won_count": data["won_count"],
        "lost_count": data["lost_count"],
        "conversion": conversion,
        "by_user": data["rows"],
    })

@bp_amocrm_company_api.get("/<int:company_id>/crm/stats.xlsx")
def crm_stats_xlsx(company_id: int):
    tok = _refresh_if_needed(company_id)
    if not tok:
        return jsonify({"error": "not connected"}), 400

    base_domain = tok.get("base_domain")
    access_token = tok.get("access_token")
    if not base_domain or not access_token:
        return jsonify({"error": "not connected"}), 400

    ts_from, ts_to, days, label = _period_from_request()
    base = _compute_stats(base_domain, access_token, ts_from, ts_to)

    sort = request.args.get("sort", "won_desc")
    try:
        min_total = int(request.args.get("min_total", "0"))
    except Exception:
        min_total = 0
    q = request.args.get("q", "").strip()

    rows = _apply_view_filters(list(base["rows"]), sort, min_total, q)
    total_won = sum(r["won"] for r in rows)
    total_lost = sum(r["lost"] for r in rows)
    total_all = total_won + total_lost
    conv_overall = round(100 * total_won / total_all) if total_all else 0

    filename_label = "today" if label == "today" else (f"{days}d" if days else "custom")
    fname = f"crm_stats_company_{company_id}_{filename_label}.xlsx"

    if Workbook is None:
        # CSV fallback c BOM (UTF-8) — Excel откроет корректно
        output = io.StringIO()
        output.write("\ufeff")
        output.write("ID;Пользователь;Успешно;Не реализовано;Конверсия (%)\n")
        for r in rows:
            output.write(f'{r["user_id"]};{r["display_name"]};{r["won"]};{r["lost"]};{r["conv"]}\n')
        mem = io.BytesIO(output.getvalue().encode("utf-8"))
        mem.seek(0)
        return send_file(
            mem,
            mimetype="text/csv; charset=utf-8",
            as_attachment=True,
            download_name=f"crm_stats_company_{company_id}_{filename_label}.csv",
        )

    # XLSX (красиво оформленный)
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика"

    ws["A1"] = "Статистика по сотрудникам"
    ws["A2"] = f"Период: {'сегодня' if label=='today' else f'последние {days} дней' if days else 'задан вручную'}"
    ws["A3"] = f"Сформировано: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())}"
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")
    ws.merge_cells("A3:E3")

    title_font = Font(size=14, bold=True)
    sub_font = Font(size=10, color="666666")
    ws["A1"].font = title_font
    ws["A2"].font = sub_font
    ws["A3"].font = sub_font

    headers = ["ID", "Пользователь", "Успешно", "Не реализовано", "Конверсия"]
    start_row = 5
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="6D5CF7")
        cell.border = Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            top=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD"),
        )

    zebra = PatternFill("solid", fgColor="F7F8FC")
    for i, r in enumerate(rows, start=1):
        row_i = start_row + i
        ws.cell(row=row_i, column=1, value=r["user_id"])
        ws.cell(row=row_i, column=2, value=r["display_name"])
        ws.cell(row=row_i, column=3, value=r["won"])
        ws.cell(row=row_i, column=4, value=r["lost"])
        conv_cell = ws.cell(row=row_i, column=5, value=(r["conv"] / 100))
        conv_cell.number_format = "0%"
        for c in range(1, 6):
            cell = ws.cell(row=row_i, column=c)
            cell.border = Border(
                left=Side(style="thin", color="EEEEEE"),
                right=Side(style="thin", color="EEEEEE"),
                top=Side(style="thin", color="EEEEEE"),
                bottom=Side(style="thin", color="EEEEEE"),
            )
            if i % 2 == 0:
                cell.fill = zebra

    total_row = start_row + len(rows) + 2
    ws.cell(row=total_row, column=1, value="ИТОГО:").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=total_won).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=total_lost).font = Font(bold=True)
    c_over = ws.cell(row=total_row, column=5, value=(conv_overall / 100))
    c_over.number_format = "0%"
    c_over.font = Font(bold=True)

    ws.freeze_panes = ws["A6"]
    for col in range(1, 6):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[chr(64 + col)].width = min(max_len + 2, 50)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )

@bp_amocrm_company_api.get("/<int:company_id>/crm/users")
def crm_users(company_id: int):
    tok = _refresh_if_needed(company_id)
    if not tok:
        return jsonify({"connected": False})
    users = []
    try:
        base_domain = tok.get("base_domain")
        url = f"https://{base_domain}/api/v4/users"
        r = requests.get(url, headers=_amo_headers(tok["access_token"]), timeout=15)
        if r.status_code == 200:
            data = r.json()
            for u in data.get("_embedded", {}).get("users", []):
                users.append({"id": u.get("id"), "name": u.get("name"), "email": u.get("email")})
    except requests.RequestException:
        current_app.logger.warning("AMO users fetch failed")
    return jsonify({"connected": True, "users": users})

@bp_amocrm_company_api.get("/<int:company_id>/crm/map/list")
def crm_map_list(company_id: int):
    m = _get_mapping(company_id)
    return jsonify({"map": m, "count": len(m)})

@bp_amocrm_company_api.post("/<int:company_id>/crm/map")
def crm_map(company_id: int):
    body = request.get_json(silent=True) or {}
    platform_id = body.get("platform_user_id")
    amocrm_user_id = body.get("amocrm_user_id")
    if platform_id is None or amocrm_user_id is None:
        return jsonify({"error": "platform_user_id and amocrm_user_id required"}), 400

    _set_mapping(company_id, int(platform_id), int(amocrm_user_id))
    return jsonify({"ok": True})

# ===== Реалтайм «сегодня» для дэшборда =====
@bp_amocrm_company_api.get("/<int:company_id>/crm/rt")
def crm_realtime(company_id: int):
    tok = _refresh_if_needed(company_id)
    if not tok:
        return jsonify({"error": "not connected"}), 400

    base_domain = tok.get("base_domain")
    access_token = tok.get("access_token")
    if not base_domain or not access_token:
        return jsonify({"error": "not connected"}), 400

    now = int(time.time())
    lt = time.localtime(now)
    midnight = int(time.mktime((lt.tm_year, lt.tm_mon, lt.tm_mday, 0, 0, 0, lt.tm_wday, lt.tm_yday, lt.tm_isdst)))

    won_by = defaultdict(int)
    lost_by = defaultdict(int)
    for lead in _iter_closed_leads(base_domain, access_token, midnight, now):
        uid = lead.get("responsible_user_id") or 0
        sid = int(lead.get("status_id") or 0)
        if sid == WON_STATUS_ID:
            won_by[uid] += 1
        elif sid == LOST_STATUS_ID:
            lost_by[uid] += 1

    created_by = defaultdict(int)
    for lead in _iter_created_leads(base_domain, access_token, midnight, now):
        uid = lead.get("responsible_user_id") or 0
        created_by[uid] += 1

    users_map = _fetch_users_map(base_domain, access_token)  # id->{name,email}

    total_won = sum(won_by.values())
    total_lost = sum(lost_by.values())
    total_created = sum(created_by.values())
    total_all = total_won + total_lost
    conversion = round(100 * total_won / total_all) if total_all else 0

    hours_passed = max(1.0, (now - midnight) / 3600.0)

    per_user = []
    user_ids = set(list(won_by.keys()) + list(lost_by.keys()) + list(created_by.keys()))
    for uid in user_ids:
        w = int(won_by.get(uid, 0))
        l = int(lost_by.get(uid, 0))
        c = int(created_by.get(uid, 0))
        tot = w + l
        conv = round(100 * w / tot) if tot else 0
        vph = round(w / hours_passed, 2)
        meta = users_map.get(int(uid), {})
        per_user.append({
            "amocrm_user_id": int(uid),
            "amocrm_name": meta.get("name") or f"User {uid}",
            "amocrm_email": meta.get("email") or "",
            "won": w,
            "lost": l,
            "created": c,
            "conv": conv,
            "wins_per_hour": vph,
        })

    per_user.sort(key=lambda r: (-r["won"], -r["conv"], -r["created"]))

    return jsonify({
        "range": "today",
        "from": midnight,
        "to": now,
        "updated_at": now,
        "kpi": {
            "won_today": total_won,
            "lost_today": total_lost,
            "created_today": total_created,
            "conversion_today": conversion,
        },
        "users": per_user,
    })


# ===== Pages =====
@bp_amocrm_pages.get("/<int:company_id>/crm")
def company_crm_page(company_id: int):
    slug = None
    if Company and db:
        c = _get_company(company_id)
        slug = getattr(c, "slug", None) if c else None
    return render_template("partner_company_crm.html", company_id=company_id, company_slug=slug or "")

@bp_amocrm_pages.get("/<int:company_id>/crm/dashboard")
def company_crm_dashboard(company_id: int):
    slug = None
    if Company and db:
        c = _get_company(company_id)
        slug = getattr(c, "slug", None) if c else None
    return render_template("partner_company_crm_dashboard.html", company_id=company_id, company_slug=slug or "")

@bp_amocrm_pages.get("/<int:company_id>")
def company_page_alias(company_id: int):
    return redirect(f"/partner/company/{company_id}/crm", code=302)

bp_amocrm_company = bp_amocrm_company_api
bp_amocrm = bp_amocrm_company_api
